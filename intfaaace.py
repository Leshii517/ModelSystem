from PyQt5.QtWidgets import QWidget, QApplication, QPushButton, QLabel, QHBoxLayout, QVBoxLayout, QLineEdit, QComboBox
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem
from random import choice
from pylab import *
import random_values
from math import *
import openpyxl
from ECM import ECM
from PRINTER import *
import random


class MainWindow(QWidget):
    def __init__(self):
        # Вызвается метод суперкласса
        super().__init__()
        # Запускается метод, в котором прописан интерфейс
        self.interface()



    def interface(self):
        '''Создаёт интерфейс'''
        #self.setWindowFlags(Qt.MSWindowsFixedSizeDialogHint)
        self.resize(1000, 950)
        #self.setWindowIcon(QIcon('icon.jpg'))
        self.setWindowTitle('SMO')

        self.shell = QVBoxLayout()

        self.main_box = QHBoxLayout()
        self.shell.addLayout(self.main_box)

        self.modeling_setting_box = QVBoxLayout()
        self.main_box.addLayout(self.modeling_setting_box)

        self.det_model_info = QLabel('Настройки детерминированной модели СМО')
        self.det_model_info.setMaximumHeight(10)
        self.modeling_setting_box.addWidget(self.det_model_info)

        self.ecm_count_box = QHBoxLayout()
        self.ecm_count_info = QLabel('Количество ЭВМ:')
        self.ecm_count_cnt = QLineEdit('10')
        self.modeling_setting_box.addLayout(self.ecm_count_box)
        self.ecm_count_box.addWidget(self.ecm_count_info)
        self.ecm_count_box.addWidget(self.ecm_count_cnt)

        self.printer_count_box = QHBoxLayout()
        self.printer_count_info = QLabel('Количество принтеров:')
        self.printer_count_cnt = QLineEdit('5')
        self.modeling_setting_box.addLayout(self.printer_count_box)
        self.printer_count_box.addWidget(self.printer_count_info)
        self.printer_count_box.addWidget(self.printer_count_cnt)

        self.user_interval_box = QHBoxLayout()
        self.user_interval_info = QLabel('Интервал прихода пользователя:')
        self.user_interval_cnt = QLineEdit('10')
        self.user_interval_box.addWidget(self.user_interval_info)
        self.user_interval_box.addWidget(self.user_interval_cnt)
        self.modeling_setting_box.addLayout(self.user_interval_box)

        self.user_task_interval_box = QHBoxLayout()
        self.user_task_interval_info = QLabel('Время работы за ЭВМ:')
        self.user_task_interval_cnt = QLineEdit('15')
        self.user_task_interval_box.addWidget(self.user_task_interval_info)
        self.user_task_interval_box.addWidget(self.user_task_interval_cnt)
        self.modeling_setting_box.addLayout(self.user_task_interval_box)

        self.printer_use_time_box = QHBoxLayout()
        self.printer_use_time_info = QLabel('Время работы за принтером:')
        self.printer_use_time_cnt = QLineEdit('5')
        self.printer_use_time_box.addWidget(self.printer_use_time_info)
        self.printer_use_time_box.addWidget(self.printer_use_time_cnt)
        self.modeling_setting_box.addLayout(self.printer_use_time_box)

        self.use_printer_chance_box = QHBoxLayout()
        self.use_printer_chance_info = QLabel('Вероятность использования принтера:')
        self.use_printer_chance_cnt = QLineEdit('30')
        self.use_printer_chance_box.addWidget(self.use_printer_chance_info)
        self.use_printer_chance_box.addWidget(self.use_printer_chance_cnt)
        self.modeling_setting_box.addLayout(self.use_printer_chance_box)

        self.model_time_box = QHBoxLayout()
        self.modeling_setting_box.addLayout(self.model_time_box)
        self.model_time_info = QLabel('Время моделирования:')
        self.model_time_box.addWidget(self.model_time_info)
        self.model_time_cnt = QLineEdit('10080')
        self.model_time_box.addWidget(self.model_time_cnt)

        self.imitation_model_info = QLabel('Настройки имитационной модели СМО')
        self.imitation_model_info.setMaximumHeight(10)
        self.modeling_setting_box.addWidget(self.imitation_model_info)

        self.user_task_interval_info = QLabel('Параметры для времени решения задачи пользователя')
        self.user_task_interval_info.setMaximumHeight(20)
        self.modeling_setting_box.addWidget(self.user_task_interval_info)
        self.user_task_interval_Hbox = QHBoxLayout()
        self.modeling_setting_box.addLayout(self.user_task_interval_Hbox)
        self.user_task_interval_m_info = QLabel('m:')
        self.user_task_interval_m_cnt = QLineEdit('15')
        self.user_task_interval_sigma_info = QLabel('sigma:')
        self.user_task_interval_sigma_cnt = QLineEdit('3')
        self.user_task_interval_Hbox.addWidget(self.user_task_interval_m_info)
        self.user_task_interval_Hbox.addWidget(self.user_task_interval_m_cnt)
        self.user_task_interval_Hbox.addWidget(self.user_task_interval_sigma_info)
        self.user_task_interval_Hbox.addWidget(self.user_task_interval_sigma_cnt)

        self.printer_use_time_info = QLabel('Параметры для времени вывода принтера')
        self.printer_use_time_info.setMaximumHeight(20)
        self.modeling_setting_box.addWidget(self.printer_use_time_info)
        self.printer_use_time_Hbox = QHBoxLayout()
        self.modeling_setting_box.addLayout(self.printer_use_time_Hbox)
        self.printer_use_time_m_info = QLabel('m:')
        self.printer_use_time_m_cnt = QLineEdit('5')
        self.printer_use_time_sigma_info = QLabel('sigma:')
        self.printer_use_time_sigma_cnt = QLineEdit('1')
        self.printer_use_time_Hbox.addWidget(self.printer_use_time_m_info)
        self.printer_use_time_Hbox.addWidget(self.printer_use_time_m_cnt)
        self.printer_use_time_Hbox.addWidget(self.printer_use_time_sigma_info)
        self.printer_use_time_Hbox.addWidget(self.printer_use_time_sigma_cnt)

        self.user_interval_info = QLabel('Параметр lambda для времени между появлением пользователей')
        self.modeling_setting_box.addWidget(self.user_interval_info)
        self.user_interval_lambda_box = QHBoxLayout()
        self.modeling_setting_box.addLayout(self.user_interval_lambda_box)
        self.user_interval_lambda_info = QLabel('lambda:')
        self.user_interval_lambda_box.addWidget(self.user_interval_lambda_info)
        self.user_interval_lambda_cnt = QLineEdit('0.1')
        self.user_interval_lambda_box.addWidget(self.user_interval_lambda_cnt)

        self.use_printer_chance_info = QLabel('Интервал разброса вероятности использования принтера')
        self.modeling_setting_box.addWidget(self.use_printer_chance_info)
        self.use_printer_chance_box = QHBoxLayout()
        self.use_printer_chance_min_info = QLabel('min:')
        self.use_printer_chance_box.addWidget(self.use_printer_chance_min_info)
        self.use_printer_chance_min_cnt = QLineEdit('5')
        self.use_printer_chance_box.addWidget(self.use_printer_chance_min_cnt)
        self.use_printer_chance_max_info = QLabel('max:')
        self.use_printer_chance_box.addWidget(self.use_printer_chance_max_info)
        self.use_printer_chance_max_cnt = QLineEdit('95')
        self.use_printer_chance_box.addWidget(self.use_printer_chance_max_cnt)
        self.modeling_setting_box.addLayout(self.use_printer_chance_box)

        self.model_type_switch = QComboBox(self)
        self.model_type_switch.addItems(
            ['Детерминированная модель', 'Имитационная модель', 'Имитационная модель с возмущением'])
        self.modeling_setting_box.addWidget(self.model_type_switch)

        self.statistics_box = QVBoxLayout()
        self.modeling_setting_box.addLayout(self.statistics_box)

        self.middle_line_length_ecm = QLabel('Средняя длина очереди на ЭВМ: ')
        self.statistics_box.addWidget(self.middle_line_length_ecm)

        self.middle_line_length_printers = QLabel('Средняя длина очереди на принтеры: ')
        self.statistics_box.addWidget(self.middle_line_length_printers)

        self.load_factor_ecm = QLabel('Коэффициент загрузки ЭВМ: ')
        self.statistics_box.addWidget(self.load_factor_ecm)

        self.load_factor_printers = QLabel('Коэффициент загрузки принтеров: ')
        self.statistics_box.addWidget(self.load_factor_printers)

        self.button_modeling = QPushButton('Моделирование')
        self.modeling_setting_box.addWidget(self.button_modeling)
        self.button_modeling.clicked.connect(self.computing_device_modeling)

        self.random_test_box = QVBoxLayout()

        self.random_test_settings_box = QHBoxLayout()
        self.random_test_m_info = QLabel('m:')
        self.random_test_m_cnt = QLineEdit('15')
        self.random_test_sigma_info = QLabel('sigma:')
        self.random_test_sigma_cnt = QLineEdit('3')
        self.random_test_lambda_info = QLabel('lambda:')
        self.random_test_lambda_cnt = QLineEdit('0.1')
        self.random_test_settings_box.addWidget(self.random_test_m_info)
        self.random_test_settings_box.addWidget(self.random_test_m_cnt)
        self.random_test_settings_box.addWidget(self.random_test_sigma_info)
        self.random_test_settings_box.addWidget(self.random_test_sigma_cnt)
        self.random_test_settings_box.addWidget(self.random_test_lambda_info)
        self.random_test_settings_box.addWidget(self.random_test_lambda_cnt)
        self.random_test_box.addLayout(self.random_test_settings_box)

        self.graph_switch_box = QHBoxLayout()
        self.distribution_switch = QComboBox()
        self.distribution_switch.addItems(['Нормальное распределение', 'Экспоненциальное распределение'])
        self.graph_switch_box.addWidget(self.distribution_switch)
        self.random_method_switch = QComboBox()
        self.random_method_switch.addItems(
            ['Метод серединных квадратов', 'Метод иррационального числа', 'Конгруэнтный метод'])
        self.graph_switch_box.addWidget(self.random_method_switch)
        self.random_test_box.addLayout(self.graph_switch_box)

        self.figure_1 = plt.figure(facecolor='0.945')
        self.chart_1 = FigureCanvas(self.figure_1)
        self.random_test_box.addWidget(self.chart_1)

        self.SKO = QLabel('Среднеквадратичное отклонение:')
        self.random_test_box.addWidget(self.SKO)

        self.figure_2 = plt.figure(facecolor='0.945')
        self.chart_2 = FigureCanvas(self.figure_2)
        self.random_test_box.addWidget(self.chart_2)

        self.kolmogorov_hypothesis = QLabel('Критерий согласия Колмогорова: ')
        self.random_test_box.addWidget(self.kolmogorov_hypothesis)

        self.pirson_hypothesis = QLabel('Критерий согласия Пирсона: ')
        self.random_test_box.addWidget(self.pirson_hypothesis)

        self.korel_moment = QLabel('Корелляционный метод: ')
        self.random_test_box.addWidget(self.korel_moment)

        self.button_test_random = QPushButton('Тест')
        self.random_test_box.addWidget(self.button_test_random)

        self.main_box.addLayout(self.random_test_box)

        self.disturbing_proc_box = QVBoxLayout()

        self.disturb_data_Hbox = QHBoxLayout()
        self.ymin_info = QLabel('ymin:')
        self.disturb_data_Hbox.addWidget(self.ymin_info)
        self.ymin_cnt = QLineEdit('5')
        self.disturb_data_Hbox.addWidget(self.ymin_cnt)
        self.ymax_info = QLabel('ymax:')
        self.disturb_data_Hbox.addWidget(self.ymax_info)
        self.ymax_cnt = QLineEdit('95')
        self.disturb_data_Hbox.addWidget(self.ymax_cnt)
        self.disturbing_proc_box.addLayout(self.disturb_data_Hbox)

        self.figure_3 = plt.figure(facecolor='0.945')
        self.chart_3 = FigureCanvas(self.figure_3)
        self.disturbing_proc_box.addWidget(self.chart_3)

        self.figure_4 = plt.figure(facecolor='0.945')
        self.chart_4 = FigureCanvas(self.figure_4)
        self.disturbing_proc_box.addWidget(self.chart_4)

        self.figure_5 = plt.figure(facecolor='0.945')
        self.chart_5 = FigureCanvas(self.figure_5)
        self.disturbing_proc_box.addWidget(self.chart_5)

        self.figure_6 = plt.figure(facecolor='0.945')
        self.chart_6 = FigureCanvas(self.figure_6)
        self.disturbing_proc_box.addWidget(self.chart_6)

        self.figure_7 = plt.figure(facecolor='0.945')
        self.chart_7 = FigureCanvas(self.figure_7)
        self.disturbing_proc_box.addWidget(self.chart_7)

        self.smirnov_info = QLabel('Критерий согласия Смирнова')
        self.disturbing_proc_box.addWidget(self.smirnov_info)
        self.smirnov_info_output = QLabel('Cтатистическая гипотеза верна с вероятностью не менее, чем ')
        self.disturbing_proc_box.addWidget(self.smirnov_info_output)

        self.stud_info = QLabel('Критерий согласия Стьюдента')
        self.disturbing_proc_box.addWidget(self.stud_info)
        self.stud_info_output = QLabel('Cтатистическая гипотеза верна с вероятностью не менее, чем ')
        self.disturbing_proc_box.addWidget(self.stud_info_output)

        self.fisher_info = QLabel('Критерий согласия Фишера')
        self.disturbing_proc_box.addWidget(self.fisher_info)
        self.fisher_info_output = QLabel('Cтатистическая гипотеза верна с вероятностью не менее, чем ')
        self.disturbing_proc_box.addWidget(self.fisher_info_output)

        self.disturbing_button = QPushButton('Тест')
        self.disturbing_button.clicked.connect(self.disturbing_proc_generate)
        self.disturbing_proc_box.addWidget(self.disturbing_button)

        self.disturbing_proc_generate()

        self.main_box.addLayout(self.disturbing_proc_box)

        self.button_test_random.clicked.connect(self.test_distribution)

        self.setLayout(self.shell)

        self.test_distribution()

        self.show()

    def read_start_values(self):
        global ecm_users
        ecm_users = []
        global printer_users
        printer_users = []
        global ecm_pending_users
        ecm_pending_users = []
        global printer_pending_users
        printer_pending_users = []
        global ecm_list
        ecm_list = []
        global printer_list
        printer_list = []
        global ecm_count
        ecm_count = int(self.ecm_count_cnt.text())
        global printer_count
        printer_count = int(self.printer_count_cnt.text())
        global user_interval
        user_interval = int(self.user_interval_cnt.text())
        global user_task_interval
        user_task_interval = int(self.user_task_interval_cnt.text())
        global printer_use_time
        printer_use_time = int(self.printer_use_time_cnt.text())
        global use_printer_chance
        use_printer_chance = int(self.use_printer_chance_cnt.text())
        global model_time
        model_time = int(self.model_time_cnt.text())

        for i in range(ecm_count):
            ecm_list.append(ECM(i))  # создаем ЭВМ
        for i in range(printer_count):
            printer_list.append(printer(i))  # создаем принтеры

        global user_task_interval_m
        user_task_interval_m = float(self.user_task_interval_m_cnt.text())
        global user_task_interval_sigma
        user_task_interval_sigma = float(self.user_task_interval_sigma_cnt.text())
        global printer_use_time_m
        printer_use_time_m = float(self.printer_use_time_m_cnt.text())
        global printer_use_time_sigma
        printer_use_time_sigma = float(self.printer_use_time_sigma_cnt.text())
        global user_interval_lambda
        user_interval_lambda = float(self.user_interval_lambda_cnt.text())
        global use_printer_chance_min
        use_printer_chance_min = int(self.use_printer_chance_min_cnt.text())
        global use_printer_chance_max
        use_printer_chance_max = int(self.use_printer_chance_max_cnt.text())
        self.ymin = int(self.ymin_cnt.text())
        self.ymax = int(self.ymax_cnt.text())

    def max_list_el(self, lst):
        max = lst[0]
        for el in lst:
            if el > max:
                max = el
        return max

    def min_list_el(self, lst):
        min = lst[0]
        for el in lst:
            if el < min:
                min = el
        return min

    def build_graph_gistogramm(self, switch, coord_x=None, coord_y=None, color=None):
        if switch == 'create':
            self.figure_1.clear()
            self.bar_chart = self.figure_1.add_subplot(111)
            self.bar_chart.grid(True)
        elif switch == 'build':
            self.bar_chart.plot(coord_x, coord_y, color)
            self.chart_1.draw()

    def build_graph_acceptance_criterion(self, switch, coord_x=None, coord_y=None, color=None):
        if switch == 'create':
            self.figure_2.clear()
            self.acceptance_criterion_chart = self.figure_2.add_subplot(111)
            self.acceptance_criterion_chart.grid(True)
        elif switch == 'build':
            self.acceptance_criterion_chart.plot(coord_x, coord_y, color)
            self.chart_2.draw()

    def build_graph(self, switch, figure, chart, coord_x=None, coord_y=None, color=None):
        if switch == 'create':
            figure.clear()
            self.correl_criterion_chart = figure.add_subplot(111)
            self.correl_criterion_chart.grid(True)
        elif switch == 'build':
            self.correl_criterion_chart.plot(coord_x, coord_y, color)
            chart.draw()

    def create_static_row(self, m, sigma, lmbd):
        static_row = []
        n = 40
        N = n * 1000
        if self.random_method_switch.currentText() == 'Метод серединных квадратов':
            x1 = random_values.random_mid_sqr(N)
            x2 = random_values.random_irrat_num(N)
        elif self.random_method_switch.currentText() == 'Метод иррационального числа':
            x1 = random_values.random_irrat_num(N)
            x2 = random_values.random_mid_sqr(N)
        else:
            x1 = random_values.random_cong(N)
            x2 = random_values.random_irrat_num(N)

        if self.distribution_switch.currentText() == 'Нормальное распределение':
            for i in range(N):
                static_row.append(random_values.normal_distribution(x1[i], x2[i], m, sigma))
        else:
            for i in range(N):
                static_row.append(random_values.exponential_distribution(x1[i], lmbd))
        return static_row

    def test_distribution(self):
        self.build_graph_gistogramm('create')
        #self.build_graph_acceptance_criterion('create')
        self.build_graph_acceptance_criterion('create')

        coord_x_ideal, coord_y_ideal = [], []
        m = float(self.random_test_m_cnt.text())
        sigma = float(self.random_test_sigma_cnt.text())
        lmbd = float(self.random_test_lambda_cnt.text())
        n = 40
        N = n * 1000
        static_row = self.create_static_row(m, sigma, lmbd)

        max = self.max_list_el(static_row)
        min = self.min_list_el(static_row)
        step = (max - min) / n # n

        dx = 0
        if self.distribution_switch.currentText() == 'Нормальное распределение':
            for i in range(n):
                coord_x_ideal.append(dx)
                coord_y_ideal.append(random_values.normal_ideal(dx, m, sigma))
                dx += step
            r = 2
        else:
            for i in range(n):
                coord_x_ideal.append(dx)
                coord_y_ideal.append(random_values.exponential_ideal(dx, lmbd))
                dx += step
            r = 3

        self.build_graph_gistogramm('build', coord_x_ideal, coord_y_ideal, 'r')

        intervals = []
        for i in range(n):
            intervals.append([min, min + step, 0])
            min += step

        for el in static_row:
            for interval in intervals:
                if el >= interval[0] and el <= interval[1]:
                    interval[2] += 1
                    break

        for interval in intervals:
            interval[2] /= N
            interval[2] /= step

        # Колмогоров
        coord_x, coord_y = [], []
        dx, dy = 0, 0
        Fcm = []
        for interval in intervals:
            coord_x.append(dx)
            dy += interval[2]
            Fcm.append(dy)
            coord_y.append(dy)
            coord_y.append(dy)
            dx += step
            coord_y.append(0)
            coord_x.append(dx)
            coord_x.append(dx)

        self.build_graph_acceptance_criterion('build', coord_x, coord_y, 'b')

        coord_x, coord_y = [], []
        dx, dy = 0, 0
        for i in range(n):
            if self.distribution_switch.currentText() == 'Экспоненциальное распределение':
                dy += coord_y_ideal[i] * 0.86
            else:
                dy += coord_y_ideal[i]
            Fcm[i] = abs(Fcm[i] - dy)
            coord_x.append(dx)
            coord_y.append(dy)
            dx += step

        D = self.max_list_el(Fcm)
        lmbd_Kolmogorov = D * sqrt(N)
        p = 1
        for k in range(-10000, 10000, 1):
            p -= (((-1) ** k) * exp(-2 * (k ** 2) * (lmbd_Kolmogorov ** 2)))
        self.kolmogorov_hypothesis.setText('Критерий согласия Колмогорова: %s' % (str((1 - p) * 100)) + '%')

        # Пирсон
        pirs = 1
        for i in range(n):
            raz = (intervals[i][2] * step) - coord_y_ideal[i]
            raz **= 2
            raz /= coord_y_ideal[i]
            pirs += raz
        pirs *= N
        pirs = float(str(sqrt(pirs)/10)[:str(sqrt(pirs)).find('e')])

        p_tabl = [99, 98, 95, 90, 80, 70, 50, 30]

        pirson_tabl = [[0.0002, 0.001, 0.004, 0.016, 0.064, 0.148, 0.455, 1.07],
        [0.020, 0.040, 0.103, 0.211, 0.446, 0.71, 1.39, 2.41],
        [0.115, 0.185, 0.352, 0.584, 1.01, 1.42, 2.37, 3.66],
        [0.297, 0.429, 0.711, 1.06, 1.65, 2.19, 3.36, 4.88],
        [0.554, 0.752, 1.15, 1.61, 2.34, 3.00, 4.35, 6.06],
        [0.872, 1.13, 1.64, 2.20, 3.07, 3.83, 5.35, 7.23],
        [1.24, 1.56, 2.17, 2.83, 3.82, 4.67, 6.35, 8.38],
        [1.65, 2.03, 2.73, 3.49, 4.59, 5.53, 7.34, 9.52],
        [2.09, 2.53, 3.33, 4.17, 5.38, 6.39, 8.34, 10.66],
        [2.56, 3.06, 3.94, 4.87, 6.18, 7.27, 9.34, 11.78],
        [3.05, 3.61, 4.57, 5.58, 6.99, 8.15, 10.34, 12.90],
        [3.57, 4.18, 5.23, 6.30, 7.81, 9.03, 11.34, 14.01],
        [4.11, 4.77, 5.89, 7.04, 8.63, 9.93, 12.34, 15.12],
        [4.66, 5.37, 6.57, 7.79, 9.47, 10.82, 13.34, 16.22],
        [5.23, 5.98, 7.26, 8.55, 10.31, 11.72, 14.34, 17.32],
        [5.81, 6.61, 7.96, 9.31, 11.15, 12.62, 15.34, 18.42],
        [6.41, 7.26, 8.67, 10.09, 12.00, 13.53, 16.34, 19.51],
        [7.01, 7.91, 9.39, 10.86, 12.86, 14.44, 17.34, 20.60],
        [7.63, 8.57, 10.12, 11.65, 13.72, 15.35, 18.34, 21.69],
        [8.26, 9.24, 10.85, 12.44, 14.58, 16.27, 19.34, 22.77],
        [8.90, 9.91, 11.59, 13.24, 15.44, 17.18, 20.34, 23.86],
        [9.54, 10.60, 12.34, 14.04, 16.31, 18.10, 21.34, 24.94],
        [10.20, 11.29, 13.09, 14.85, 17.19, 19.02, 22.34, 26.02],
        [10.86, 11.99, 13.85, 15.66, 18.06, 19.94, 23.34, 27.10],
        [11.52, 12.70, 14.61, 16.47, 18.94, 20.87, 24.34, 28.17],
        [12.20, 13.41, 15.38, 17.29, 19.82, 21.79, 25.34, 29.25],
        [12.88, 14.13, 16.15, 18.11, 20.70, 22.72, 26.34, 30.32],
        [13.56, 14.85, 16.93, 18.94, 21.59, 23.65, 27.34, 31.39],
        [14.26, 15.57, 17.71, 19.77, 22.48, 24.58, 28.34, 32.46],
        [14.95, 16.31, 18.49, 20.60, 23.36, 25.51, 29.34, 33.53],
        [15.66, 17.04, 19.28, 21.43, 24.26, 26.44, 30.34, 34.60],
        [16.36, 17.78, 20.07, 22.27, 25.15, 27.37, 31.34, 35.66],
        [17.07, 18.53, 20.87, 23.11, 26.04, 28.31, 32.34, 36.73],
        [17.79, 19.28, 21.66, 23.95, 26.94, 29.24, 33.34, 37.80],
        [18.51, 20.03, 22.47, 24.80, 27.84, 30.18, 34.34, 38.86],
        [19.23, 20.78, 23.27, 25.64, 28.73, 31.12, 35.34, 39.92],
        [19.96, 21.54, 24.07, 26.49, 29.64, 32.05, 36.34, 40.98],
        [20.69, 22.30, 24.88, 27.34, 30.54, 32.99, 37.34, 42.05],
        [21.43, 23.07, 25.70, 28.20, 31.44, 33.93, 38.34, 43.11],
        [22.16, 23.84, 26.51, 29.05, 32.34, 34.87, 39.34, 44.16],
        [22.91, 24.61, 27.33, 29.91, 33.25, 35.81, 40.34, 45.22],
        [23.65, 25.38, 28.14, 30.77, 34.16, 36.75, 41.34, 46.28],
        [24.40, 26.16, 28.96, 31.63, 35.07, 37.70, 42.34, 47.34],
        [25.15, 26.94, 29.79, 32.49, 35.97, 38.64, 43.34, 48.40],
        [25.90, 27.72, 30.61, 33.35, 36.88, 39.58, 44.34, 49.45],
        [26.66, 28.50, 31.44, 34.22, 37.80, 40.53, 45.34, 50.5],
        [27.42, 29.29, 32.27, 35.08, 38.71, 41.47, 46.34, 51.56]]

        for i in range(7, -1, -1):
            if pirs < pirson_tabl[n - r - 1][i]:
                break
        self.pirson_hypothesis.setText('Критерий согласия Пирсона: %s' % (str(p_tabl[i]) + '%'))

        # Корреляционный момент
        t = 1
        summ_1, summ_2, summ_3, summ_4, summ_5 = 0, 0, 0, 0, 0
        Kxy, Dx, Dy = 0, 0, 0
        minp = 0
        for i in range(N - t):
            summ_1 += static_row[i] * static_row[i + t]
            summ_2 += static_row[i]
            summ_3 += static_row[i + t]
            summ_4 += static_row[i] ** 2
            summ_5 += static_row[i + t] ** 2

        Kxy = (summ_1 / (N - t)) - (summ_2 * summ_3 / ((N - t) ** 2))
        Dx = (summ_4 / (N - t)) - ((summ_2 / (N - t)) ** 2)
        Dy = (summ_5 / (N - t)) - ((summ_3 / (N - t)) ** 2)
        pxy = abs(Kxy / sqrt(Dx * Dy))
        self.korel_moment.setText('Корелляционный метод: %s' % (str((1 - pxy) * 100) + '%'))

        # Гистограмма
        self.build_graph_acceptance_criterion('build', coord_x, coord_y, 'r')

        coord_x, coord_y = [], []

        dx = 0

        for i in range(n):
            coord_x.append(dx)
            coord_y.append(intervals[i][2])
            coord_y.append(intervals[i][2])
            dx += step
            coord_y.append(0)
            coord_x.append(dx)
            coord_x.append(dx)

        self.build_graph_gistogramm('build', coord_x, coord_y, 'b')

        sko = 0
        for i in range(len(coord_y_ideal)):
            sko += (coord_y_ideal[i] - intervals[i][2]) ** 2
        sko /= len(intervals)
        sko = sqrt(sko)

        self.SKO.setText('Среднеквадратичное отклонение: %s' % str(sko)[:8])

    def create_gist_coords(self, static_row):
        max = self.max_list_el(static_row)
        min = self.min_list_el(static_row)
        n = 40
        N = len(static_row)
        step = (max - min) / n

        intervals = []
        for i in range(n):
            intervals.append([min, min + step, 0])
            min += step

        for el in static_row:
            for interval in intervals:
                if el >= interval[0] and el <= interval[1]:
                    interval[2] += 1
                    break

        for interval in intervals:
            interval[2] /= N
            interval[2] /= step

        coord_x, coord_y = [], []
        dx, dy = 0, 0
        Fcm = []
        for interval in intervals:
            coord_x.append(dx)
            dy += interval[2]
            Fcm.append(dy)
            coord_y.append(dy)
            coord_y.append(dy)
            dx += step
            coord_y.append(0)
            coord_x.append(dx)
            coord_x.append(dx)

        out_1 = [coord_x, coord_y]

        dx = 0
        coord_x, coord_y = [], []
        for i in range(n):
            coord_x.append(dx)
            coord_y.append(intervals[i][2])
            coord_y.append(intervals[i][2])
            dx += step
            coord_y.append(0)
            coord_x.append(dx)
            coord_x.append(dx)

        out_2 = [coord_x, coord_y]

        return out_1, out_2, Fcm

    def smirnov(self, Nx, Ny):
        self.build_graph('create', self.figure_4, self.chart_4)
        out_1, out_2, Fcmx = self.create_gist_coords(Nx)
        coord_x, coord_y = out_2
        self.build_graph('build', self.figure_4, self.chart_4, coord_x, coord_y, 'b')
        self.build_graph('create', self.figure_5, self.chart_5)
        coord_x, coord_y = out_1
        self.build_graph('build', self.figure_5, self.chart_5, coord_x, coord_y, 'b')

        self.build_graph('create', self.figure_6, self.chart_6)
        out_1, out_2, Fcmy = self.create_gist_coords(Ny)
        coord_x, coord_y = out_2
        self.build_graph('build', self.figure_6, self.chart_6, coord_x, coord_y, 'b')
        self.build_graph('create', self.figure_7, self.chart_7)
        coord_x, coord_y = out_1
        self.build_graph('build', self.figure_7, self.chart_7, coord_x, coord_y, 'b')

        D = 0
        for i in range(len(Fcmx)):
            raz = abs(Fcmx[i] - Fcmy[i])
            if raz > D:
                D = raz

        p = 0
        p_list = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 0.95, 0.98, 0.99]
        for i in range(len(p_list)):
            s = sqrt(abs(log(1 - p_list[i]) * ((1 / len(Nx)) + (1 / len(Ny))) / 2)) * 100
            if D <= s:
                p = p_list[i]

        return p

    def student(self, Nx, Ny):
        mx, my = 0, 0
        lx = len(Nx)
        ly = len(Ny)
        for i in range(lx):
            mx += Nx[i]
        for i in range(ly):
            my += Ny[i]
        mx /= lx
        my /= ly
        Dx, Dy = 0, 0
        for i in range(lx):
            Dx += ((Nx[i] - mx) ** 2) / (lx - 1)
        for i in range(ly):
            Dy += ((Ny[i] - my) ** 2) / (ly - 1)

        D = ((lx - 1) * Dx + (ly - 1) * Dy) / (lx + ly - 2)
        t = sqrt((((mx - my) ** 2) * lx * ly) / (D * (lx + ly)))

        stud_coefs = [[0.126, 0.1], [0.253, 0.2], [0.385, 0.3], [0.524, 0.4], [0.674, 0.5],
                      [0.842, 0.6] ,[1.036, 0.7], [1.282, 0.8], [1.645, 0.9], [1.960, 0.95],
                      [2.330, 0.98], [2.580, 0.99], [3.290, 0.999]]

        for i in range(len(stud_coefs)):
            if t <= stud_coefs[i][0]:
                p = stud_coefs[i][1]
        return p

    def fisher(self, Nx, Ny):
        s = 0
        for i in range(0, len(Nx)):
            s += Nx[i]
        mx = (1 / len(Nx)) * s

        dx = 0
        for i in range(0, len(Nx)):
            dx += ((Nx[i] - mx) ** 2) / (len(Nx) - 1)

        s = 0
        for i in range(0, len(Ny)):
            s += Ny[i]
        my = (1 / len(Ny)) * s

        dy = 0
        for i in range(0, len(Ny)):
            dy += ((Ny[i] - my) ** 2) / (len(Ny) - 1)

        if dx > dy:
            F = dx / dy
        else:
            F = dy / dx

        F_tabl = 1.09
        if F <= F_tabl:
            return 'Cтатистическая гипотеза верна с вероятностью не менее, чем 95%'
        else:
            return 'Cтатистическая гипотеза верна с вероятностью менее, чем 95%'

    def array_for_disturbing(self):
        self.read_start_values()
        My = (use_printer_chance_max + use_printer_chance_min) / 2
        m, sigma = 0, 1
        N = 20000
        x11 = random_values.random_cong(N + 4)
        x12 = random_values.random_cong(N + 4)
        C0, C1, C2, C3 = 4.442, 0.496, 0.003, 4.946
        Q1 = []
        Q1 = [random_values.normal_distribution(x11[i], x12[i], m, sigma) for i in range(4)]
        Y1 = []
        for i in range(4, N + 4):
            Y1.append(C0 * Q1[i - 4] + C1 * Q1[i - 3] + C2 * Q1[i - 2] + C3 * Q1[i - 1] + My)
            Q1.append(random_values.normal_distribution(x11[i], x12[i], m, sigma))
        self.Y_work = [int(Y1[i]) + 1 for i in range(len(Y1))]

    def disturbing_proc_generate(self):
        self.read_start_values()
        a = ((self.ymax - self.ymin) / 6) ** 2
        b = -log(0.05) / 3
        func_K = lambda a, b, t: a * exp(-b * 3*t)
        self.build_graph('create', self.figure_3, self.chart_3)
        coord_x, coord_y = [], []
        i = 0
        while i <= 5:
            coord_x.append(i)
            coord_y.append(func_K(a, b, i))
            i += 0.1
        self.build_graph('build', self.figure_3, self.chart_3, coord_x, coord_y, 'r')

        My = (self.ymax + self.ymin) / 2
        m, sigma = 0, 1
        N = 1000
        x1 = random_values.random_cong(N + 4)
        x2 = random_values.random_cong(N + 4)
        C0, C1, C2, C3 = 4.442, 0.496, 0.003, 4.946
        Q = []
        Q = [random_values.normal_distribution(x1[i], x2[i], m, sigma) for i in range(4)]
        Y = []

        for i in range(4, N + 4):
            Y.append(C0 * Q[i - 4] + C1 * Q[i - 3] + C2 * Q[i - 2] + C3 * Q[i - 1] + My)
            Q.append(random_values.normal_distribution(x1[i], x2[i], m, sigma))

        Nx, Ny = [], []
        for i in range(N):
            if i % 2 == 0:
                Nx.append(Y[i])
            else:
                Ny.append(Y[i])
        self.Y = [int(Y[i]) + 1 for i in range(len(Y))]
        self.stud_info_output.setText('Cтатистическая гипотеза верна с вероятностью не менее, чем %s' % str(self.student(Nx, Ny) * 100) + '%')
        self.smirnov_info_output.setText('Cтатистическая гипотеза верна с вероятностью не менее, чем %s' % str(self.smirnov(Nx, Ny) * 100) + '%')
        self.fisher_info_output.setText(self.fisher(Nx, Ny))

    def import_excel(self):
        # Вывод показателей в Excel
        # Статистика одного экспримента
        # Оптимизация
        wb = openpyxl.load_workbook('./Optim.xlsx')
        sheet = wb['Sheet1']
        sheet['A1'], sheet['B1'],  sheet['C1'] = '№', 'N', 'M'
        sheet['F1'], sheet['G1'], sheet['H1'], sheet['I1'] = 'СДЭВМ',  'СДПр', 'КЗЭВМ', 'КЗПр'
        sheet['A' + str(random_values.count_exp + 1)] = random_values.count_exp
        sheet['B' + str(random_values.count_exp + 1)] = ecm_count
        sheet['C' + str(random_values.count_exp + 1)] = printer_count

        sheet['F' + str(random_values.count_exp + 1)] = round(middle_line_length_ecm, 2)
        sheet['G' + str(random_values.count_exp + 1)] = round(middle_line_length_printers, 2)
        sheet['H' + str(random_values.count_exp + 1)] = round(load_factor_ecm, 2)
        sheet['I' + str(random_values.count_exp + 1)] = round(load_factor_printers, 2)
        wb.save('./Optim.xlsx')
        # Протокол эксель не активен, т к низкое быстродействие записи

    def computing_device_modeling(self):
        self.read_start_values()
        random_values.count_exp += 1  # количество опытов моделирования
        global middle_line_length_ecm
        middle_line_length_ecm = 0
        global middle_line_length_printers
        middle_line_length_printers = 0
        global load_factor_ecm
        load_factor_ecm = 0
        global load_factor_printers
        queue_sum_ecm = 0
        load_factor_printers = 0
        queue_sum_printer = 0
        printer_in_use_sum = 0
        ecm_in_use_sum = 0
        use_printer_chance_cnt = 0
        global user_task_interval
        global printer_use_time
        global user_interval
        global use_printer_chance
        global model_type
        value_work = [[0] * 576 for i in range(2)]   # двумерный массив для паретто, критерии
        parameters = [[0] * 576 for i in range(2)]  # двумерный массив для паретто, параметры

        if self.model_type_switch.currentText() == 'Детерминированная модель':
            model_type = 'det'
        elif self.model_type_switch.currentText() == 'Имитационная модель':
            model_type = 'imit'
        elif self.model_type_switch.currentText() == 'Имитационная модель с возмущением':
            model_type = 'imit_with_noise'

        if model_type == 'imit' or model_type == 'imit_with_noise':
            # создаем массив для генарации возмущения
            self.array_for_disturbing()
            # генерируем массивы из ПСЧ
            self.x1 = random_values.random_cong(1000)
            self.x2 = random_values.random_irrat_num(1000)
            # Определяем случайные значения параметров системы
            while user_interval == 0:
                user_interval = int(abs(random_values.exponential_distribution(choice(self.x1),
                                                                               user_interval_lambda)))

            user_task_interval = int(abs(random_values.normal_distribution(
                choice(self.x1), choice(self.x2), user_task_interval_m, user_task_interval_sigma)))
            printer_use_time = int(abs(random_values.normal_distribution(
                choice(self.x1), choice(self.x2), printer_use_time_m, printer_use_time_sigma)))

        file1 = open('info1.txt', 'w')
        file2 = open('info2.txt', 'w')

        for mode in range(0, model_time + 1):

            # задание возмущения
            if model_type == 'imit_with_noise':
                use_printer_chance = int(self.Y_work[use_printer_chance_cnt])
                use_printer_chance_cnt += 1

            if mode % user_interval == 0:  # and i / user_task_interval > 0:
                ecm_pending_users.append(User())  # в список ожидающих
                for ecm in ecm_list:  # проверяем все эвм
                    if ecm.state == 'Free':  # если эвм свободна
                        ecm_users.append(ecm_pending_users[0])  # из очереди добавляем в ЭВМ
                        ecm.state = 'Work'  # и изменяем состояние ЭВМ
                        ecm_pending_users.remove(ecm_pending_users[0])  # его же удаляем из очереди

                        ecm_users.reverse()
                        ecm_users[0].name_ecm_work = ecm.name
                        ecm_users.reverse()
                        break

            for user in ecm_users:
                user.ecm_tick()

            for user in printer_users:
                user.printer_tick()

            if len(printer_pending_users) != 0:  # если есть очередь на принтеры
                for printer in printer_list:  # проверяем все printeri
                    if printer.state == 'Free':  # если принтер свободен
                        printer_users.append(printer_pending_users[0])  # из очереди добавляем в принтеры
                        printer.state = 'Work'  # и изменяем состояние printerа
                        printer_pending_users.remove(printer_pending_users[0])  # его же удаляем из очереди

                        printer_users.reverse()
                        printer_users[0].name_printer_work = printer.name
                        printer_users.reverse()
                        break

            if len(ecm_pending_users) != 0:  # если есть очередь на ЭВМ
                for ecm in ecm_list:  # проверяем все эвм
                    if ecm.state == 'Free':  # если эвм свободна
                        ecm_users.append(ecm_pending_users[0])  # из очереди добавляем в ЭВМ
                        ecm.state = 'Work'  # и изменяем состояние ЭВМ
                        ecm_pending_users.remove(ecm_pending_users[0])  # его же удаляем из очереди

                        ecm_users.reverse()
                        ecm_users[0].name_ecm_work = ecm.name
                        ecm_users.reverse()  # User.name_ecm_work = ecm.name
                    break


            for user in ecm_users:  # Пользователи в ЭВМ
                if user.done_ecm:  # Если закончил работу с ЭВМ
                    # ecm_users.remove(user)  # Удаляем
                    if user.wants_to_use_printer == True:  # Если хочет работать с принтером
                        ecm_users.remove(user) # Удаляем
                        user.done_ecm = False
                        user.ecm_end = True
                        for ecm in ecm_list:  # цикл для нахождения освободившейся эвм
                            if user.name_ecm_work == ecm.name:
                                ecm.state = 'Free'
                                break
                        for printer in printer_list:  # цикл для нахождения свободного принтера
                            if printer.state == 'Free':  # если принтер свободен
                                printer_pending_users.append(user)
                                printer_users.append(printer_pending_users[0])  # из очереди добавляем в принтеры
                                printer.state = 'Work'  # и изменяем состояние printerа
                                printer_pending_users.remove(printer_pending_users[0])  # его же удаляем из очереди

                                printer_users.reverse()
                                printer_users[0].name_printer_work = printer.name
                                printer_users.reverse()
                                break

                            else:
                                printer_pending_users.append(user)  # если заняты принтеры в очередь
                    else:  # Если не хочет работать с принтером
                        for ecm in ecm_list:  # цикл для нахождения освободившейся эвм
                            if user.name_ecm_work == ecm.name:
                                ecm.state = 'Free'
                                break
                        ecm_users.remove(user)  # Удаляем

            for user in printer_users:
                if user.done_printer:  # закончил с принтерами
                    for printer in printer_list:  # цикл для нахождения освободившегося принтера
                        if user.name_printer_work == printer.name:
                            printer.state = 'Free'
                            break
                    printer_users.remove(user)  # Удаляем


            for ecm in ecm_list:
                if ecm.state == 'Work':
                    file1.write(str(mode) + ' ' + ecm.name + '\n')  # запись в файл

            for printer in printer_list:
                if printer.state == 'Work':
                    file2.write(str(mode) + ' ' + printer.name + '\n')  # запись в файл
            #подсчет показателей качества
            queue_sum_ecm += len(ecm_pending_users)
            queue_sum_printer += len(printer_pending_users)

            ecm_in_use_sum += len(ecm_users) / ecm_count * 100
            printer_in_use_sum += len(printer_users) / printer_count * 100



        #global middle_line_length_ecm
        middle_line_length_ecm = queue_sum_ecm / model_time
       # global middle_line_length_printers
        middle_line_length_printers = queue_sum_printer / model_time
        #global load_factor_ecm
        load_factor_ecm = ecm_in_use_sum / model_time
        #global load_factor_printers
        load_factor_printers = printer_in_use_sum / model_time

        self.middle_line_length_ecm.setText(
            'Средняя длина очереди на ЭВМ: %s' % (str(round(middle_line_length_ecm,2))))

        self.middle_line_length_printers.setText(
            'Средняя длина очереди на принтеры: %s' % (str(round(middle_line_length_printers, 2))))

        self.load_factor_ecm.setText(
            'Коэффициент загрузки ЭВМ: %s' % (str(round(load_factor_ecm, 2)) + '%'))
        self.load_factor_printers.setText(
            'Коэффициент загрузки принтеров: %s' % (str(round(load_factor_printers, 2)) + '%'))
        #self.import_excel()
        #self.pareto(parameters, value_work)
    # парето, заполняем массив из экселя
    def pareto(self, parameters, value_work):
        wb = openpyxl.load_workbook('./Optim.xlsx')
        sheet = wb['Sheet1']
        j = 0
        while j <= 575:
            value_work[0][j] = sheet['H' + str(j + 2)]
            value_work[1][j] = sheet['I' + str(j + 2)]

            parameters[0][j] = sheet['B' + str(j + 2)]
            parameters[1][j] = sheet['C' + str(j + 2)]
            j += 1
        # оптимизация
        j = 0
        while j < len(value_work[0]) - 1:
            k = 0
            while k < len(value_work[0]):
                if k != j:
                    if (value_work[0][k].value >= value_work[0][j].value) and (
                            value_work[1][k].value >= value_work[1][j].value) and (
                            parameters[0][k].value <= parameters[0][j].value)  and (
                            parameters[1][k].value <= parameters[1][j].value):
                        del value_work[0][j]
                        del value_work[1][j]
                        del parameters[0][j]
                        del parameters[1][j]
                        j -= 1
                        break
                k += 1
            j += 1
        wb.save('./Optim.xlsx')
        # вывод результатов оптимизации
        wb = openpyxl.load_workbook('./optim_result.xlsx')
        sheet = wb['Sheet1']
        sheet['A1'], sheet['B1'], sheet['C1'] = '№', 'N', 'M'
        sheet['F1'], sheet['G1'] = 'КЗЭВМ', 'КЗПр'
        j = 0
        while j < len(value_work[0]):
            sheet['A' + str(j + 2)] = j + 1
            sheet['B' + str(j + 2)] = parameters[0][j].value
            sheet['C' + str(j + 2)] = parameters[1][j].value
            sheet['F' + str(j + 2)] = value_work[0][j].value
            sheet['G' + str(j + 2)] = value_work[1][j].value

            j += 1
        wb.save('./optim_result.xlsx')

class User(MainWindow):
    def __init__(self):
        self.will_use_printer = True if (random.randint(0, 100) <= use_printer_chance) else False
        self.task_remaining_time = user_task_interval
        self.printer_remaining_time = printer_use_time
        self.wants_to_use_printer = False
        self.done_ecm = False
        self.done_printer = False
        self.name_ecm_work = ""
        self.name_printer_work = ""
        self.ecm_end = False
        '''self.will_use_printer = False
        self.task_remaining_time = 0
        self.printer_remaining_time = 0
        self.wants_to_use_printer = False
        self.done_ecm = False
        self.done_printer = False
        self.name_ecm_work = ""
        self.name_printer_work = ""
        self.ecm_end = False'''

    def ecm_tick(self):
        if self.ecm_end == False:
            if self.will_use_printer:  # если будет работать с принтером
                if self.task_remaining_time != 0:  # если еще работает с эвм
                    self.task_remaining_time -= 1  # тогда уменьшаем время до окончания
                else:
                    self.done_ecm = True  # то закончил с ЭВМ
                    self.wants_to_use_printer = True  # и ждет принтер
            else:  # если не будет работать с принтером
                if self.task_remaining_time != 0:  # если еще работает с эвм
                    self.task_remaining_time -= 1  # тогда уменьшаем время до оканчания
                else:
                    self.done_ecm = True  # иначе задание выполнено


    def printer_tick(self):
        if self.printer_remaining_time != 0:  # если работа принтера не 0
            self.printer_remaining_time -= 1  # тогда уменьшаем время до окончания принтера
        else:
            self.done_printer = True  # иначе выполнено

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MainWindow()
    sys.exit(app.exec_())